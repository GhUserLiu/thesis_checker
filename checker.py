#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
论文选题查重系统
读取"原始数据"文件夹中的Excel文件，查重"课题名称"，输出到"查重结果"文件夹
"""

import os
import sys
import glob
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import jieba
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 导入配置和工具模块
from config import (
    SIMILARITY_THRESHOLD, CSV_ENCODING, CSV_ENCODING_ERRORS,
    EXCEL_FONT_SIZE, HEADER_FONT_SIZE,
    COLOR_HEADER_BG, COLOR_HEADER_FONT,
    COLOR_HIGH_SIMILARITY, COLOR_MEDIUM_SIMILARITY,
    COLOR_INVALID_ROW, COLOR_SIMILAR_CELL,
    DEFAULT_ORGANIZATION, DEFAULT_TEMPLATE_TYPE,
    DEFAULT_SECOND_TEACHER, RANDOM_SEED
)
from utils import (
    DataCleaner, Validator, FileHelper, Logger
)

# 设置输出编码为UTF-8
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


class ThesisChecker:
    def __init__(self, data_dir="原始数据", resubmit_dir="二次提交", output_dir="查重结果", threshold=SIMILARITY_THRESHOLD):
        self.data_dir = Path(data_dir)
        self.resubmit_dir = Path(resubmit_dir)
        self.output_dir = Path(output_dir)
        self.threshold = threshold
        self.output_dir.mkdir(exist_ok=True)
        self.resubmit_dir.mkdir(exist_ok=True)  # 确保二次提交文件夹存在

    def get_excel_files(self, directory):
        """获取指定目录中的所有Excel和CSV文件"""
        excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx")) + list(directory.glob("*.csv"))
        return excel_files if excel_files else []

    def has_resubmit_data(self):
        """检查是否有二次提交数据"""
        return len(self.get_excel_files(self.resubmit_dir)) > 0

    def detect_and_read_simplified_resubmit(self, file):
        """检测并读取简化的二次提交格式

        简化格式包含：序号、班级、学生姓名、指导教师、论文题目
        返回：(是否为简化格式, DataFrame或None)
        """
        try:
            # 尝试不同的header位置
            for header_row in [None, 0, 1, 2]:
                try:
                    df = pd.read_excel(file, header=header_row)
                    cols = df.columns.tolist()

                    # 检测是否为简化格式
                    # 简化格式的特征：
                    # 1. 包含"序号"或数字编号列
                    # 2. 包含"班级"
                    # 3. 包含"学生姓名"
                    # 4. 包含"指导教师"或"导师"
                    # 5. 包含"论文题目"或"课题名称"
                    has_seq = any('序号' in str(c) or c in range(10) for c in cols)
                    has_class = any('班级' in str(c) for c in cols)
                    has_student = any('学生姓名' in str(c) or '姓名' in str(c) for c in cols)
                    has_teacher = any('指导教师' in str(c) or '导师' in str(c) for c in cols)
                    has_title = any('论文题目' in str(c) or '课题名称' in str(c) or '题目' in str(c) for c in cols)

                    if has_class and has_student and has_teacher and has_title:
                        # 确认为简化格式
                        # 标准化列名
                        column_mapping = {}

                        # 查找各列（注意顺序很重要，更具体的要放在前面）
                        for col in df.columns:
                            col_str = str(col)
                            if '序号' in col_str or col in range(10):
                                column_mapping['_序号'] = col
                            elif '班级' in col_str:
                                column_mapping['班级'] = col
                            elif '学生姓名' in col_str:
                                column_mapping['学生姓名'] = col
                            elif '姓名' in col_str:  # 简化格式可能只有"姓名"列
                                column_mapping['学生姓名'] = col
                            elif '论文题目' in col_str or '课题名称' in col_str:
                                column_mapping['课题名称'] = col
                            elif '题目' in col_str:
                                column_mapping['课题名称'] = col
                            # 指导教师姓名（必须在最后，优先匹配精确的）
                            elif col_str == '指导教师' or col_str == '导师':
                                column_mapping['指导教师姓名'] = col
                            elif '指导教师' in col_str or '导师' in col_str:
                                # 只有没有精确匹配时才使用模糊匹配
                                if '指导教师姓名' not in column_mapping:
                                    column_mapping['指导教师姓名'] = col

                        # 创建标准化的DataFrame
                        standardized_df = pd.DataFrame()
                        for std_col, orig_col in column_mapping.items():
                            # 不复制下划线开头的临时列到最终DataFrame
                            if not std_col.startswith('_'):
                                standardized_df[std_col] = df[orig_col]

                        # 添加必要字段（后续从原始数据补充）
                        standardized_df['_is_simplified_format'] = True
                        standardized_df['_original_file'] = file.name
                        # 将"班级"复制到"所属专业"以便与原始数据匹配
                        if '班级' in standardized_df.columns:
                            standardized_df['所属专业'] = standardized_df['班级'].copy()

                        return True, standardized_df
                except:
                    continue

            return False, None
        except:
            return False, None

    def read_excel_files(self):
        """读取Excel文件 - 先读取原始数据保存为中间文件,然后用二次提交更新论文题目"""
        print("\n正在读取数据文件...")

        # 创建中间数据文件夹
        intermediate_dir = Path("中间数据")
        if intermediate_dir.exists():
            # 清空中间数据文件夹
            import shutil
            shutil.rmtree(intermediate_dir)
        intermediate_dir.mkdir(parents=True, exist_ok=True)
        print("✓ 已创建/清空「中间数据」文件夹")

        # 第一步: 读取所有原始数据文件,建立基础数据库
        print("\n步骤1: 读取原始数据文件夹,建立基础数据库...")
        base_files = self.get_excel_files(self.data_dir)
        if not base_files:
            raise FileNotFoundError(f"在 {self.data_dir} 文件夹中未找到Excel文件")

        base_data = []
        invalid_data = []  # 存储无效数据

        for file in base_files:
            try:
                df = None
                # 检查是否为CSV文件
                is_csv = file.suffix.lower() == '.csv'

                if is_csv:
                    # 尝试多种编码读取CSV
                    encodings = ['gbk', 'utf-8', 'gb18030', 'utf-8-sig', 'latin1']
                    for encoding in encodings:
                        try:
                            df = pd.read_csv(file, encoding=encoding, on_bad_lines='skip')
                            break
                        except Exception:
                            continue
                else:
                    # 读取Excel文件(使用已有的自动表头检测逻辑)
                    df = self._read_excel_with_auto_header(file)

                if df is not None and len(df) > 0:
                    df_normalized = self.normalize_columns(df)
                    # normalize_columns返回 (valid_df, invalid_df)
                    valid_df, invalid_df = df_normalized
                    # 只保存有效数据作为基础数据
                    if valid_df is not None and len(valid_df) > 0:
                        valid_df['source_file'] = file.name
                        base_data.append(valid_df)
                        class_info = f" [班级: {valid_df['文件提取的班级'].iloc[0]}]" if '文件提取的班级' in valid_df.columns and valid_df['文件提取的班级'].iloc[0] else ""
                        print(f"✓ 已读取: {file.name}{class_info} ({len(valid_df)} 条记录)")
                    # 记录无效数据数量
                    if invalid_df is not None and len(invalid_df) > 0:
                        invalid_df['source_file'] = file.name
                        invalid_data.append(invalid_df)
                        print(f"  ! 跳过 {len(invalid_df)} 条无效记录")
            except Exception as e:
                print(f"✗ 读取失败: {file.name} - {e}")
                Logger.error(f"读取文件失败: {file.name} - {e}")

        # 合并所有基础数据
        if not base_data:
            raise ValueError("未能成功读取任何原始数据文件")

        import numpy as np
        base_df = pd.concat(base_data, ignore_index=True)

        # 合并所有无效数据
        base_invalid_df = pd.concat(invalid_data, ignore_index=True) if invalid_data else pd.DataFrame()

        # 确保学号和工号是文本格式
        if '学生学号' in base_df.columns:
            base_df['学生学号'] = base_df['学生学号'].astype(str)
        if '指导教师工号' in base_df.columns:
            base_df['指导教师工号'] = base_df['指导教师工号'].astype(str)

        print(f"\n✓ 基础数据库建立完成: {len(base_df)} 条记录")

        # 保存中间文件（使用Excel格式以保留文本格式的学号和工号）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        intermediate_file = intermediate_dir / f"中间数据_{timestamp}.xlsx"
        base_df.to_excel(intermediate_file, index=False, engine='openpyxl')
        print(f"✓ 中间文件已保存: {intermediate_file.name}")

        # 第二步: 如果有二次提交数据,读取并更新论文题目
        unmatched_records = []  # 存储未匹配的记录

        if self.has_resubmit_data():
            print("\n步骤2: 读取二次提交文件夹,更新论文题目...")
            resubmit_files = self.get_excel_files(self.resubmit_dir)

            # 用于存储更新的记录
            updated_count = 0
            matched_count = 0

            for file in resubmit_files:
                try:
                    df = None
                    is_csv = file.suffix.lower() == '.csv'

                    if is_csv:
                        encodings = ['gbk', 'utf-8', 'gb18030', 'utf-8-sig', 'latin1']
                        for encoding in encodings:
                            try:
                                df = pd.read_csv(file, encoding=encoding, on_bad_lines='skip')
                                break
                            except:
                                continue
                    else:
                        # 先检测是否为简化格式
                        is_simplified, df = self.detect_and_read_simplified_resubmit(file)
                        if df is None:
                            df = self._read_excel_with_auto_header(file)

                    if df is not None and len(df) > 0:
                        # 规范化二次提交数据
                        df_normalized = self.normalize_columns(df)
                        # normalize_columns返回 (valid_df, invalid_df)
                        valid_df, invalid_df = df_normalized

                        if valid_df is not None and len(valid_df) > 0:
                            # 确保学号是文本格式
                            if '学生学号' in valid_df.columns:
                                valid_df['学生学号'] = valid_df['学生学号'].astype(str)

                            # 通过"学号-学生姓名-指导教师姓名"匹配并更新论文题目
                            for idx, row in valid_df.iterrows():
                                student_id_raw = row.get('学生学号', '')
                                student_name_raw = row.get('学生姓名', '')
                                teacher_name_raw = row.get('指导教师姓名', '')
                                new_title = str(row['课题名称']).strip()

                                # 清理数据
                                student_id = str(student_id_raw).strip() if pd.notna(student_id_raw) else ''
                                student_name = str(student_name_raw).strip() if pd.notna(student_name_raw) else ''
                                teacher_name = str(teacher_name_raw).strip() if pd.notna(teacher_name_raw) else ''

                                # 跳过空值
                                if not student_name or not teacher_name:
                                    continue

                                # 在基础数据中查找匹配记录
                                match_mask = (
                                    (base_df['学生姓名'].astype(str).str.strip() == student_name) &
                                    (base_df['指导教师姓名'].astype(str).str.strip() == teacher_name)
                                )

                                # 如果有学号,也匹配学号
                                if student_id and student_id not in ['', 'nan', 'None']:
                                    match_mask = match_mask & (base_df['学生学号'].astype(str).str.strip() == student_id)

                                matches = base_df[match_mask]

                                # 如果在有效数据中没找到，尝试在无效数据中查找
                                if len(matches) == 0 and len(base_invalid_df) > 0:
                                    invalid_match_mask = (
                                        (base_invalid_df['学生姓名'].astype(str).str.strip() == student_name) &
                                        (base_invalid_df['指导教师姓名'].astype(str).str.strip() == teacher_name)
                                    )

                                    invalid_matches = base_invalid_df[invalid_match_mask]

                                    if len(invalid_matches) > 0:
                                        # 在无效数据中找到了匹配，更新课题名称并添加到基础数据
                                        for match_idx in invalid_matches.index:
                                            old_row = base_invalid_df.loc[match_idx].to_dict()
                                            old_row['课题名称'] = new_title
                                            # 添加到基础数据
                                            base_df = pd.concat([base_df, pd.DataFrame([old_row])], ignore_index=True)
                                            matched_count += 1
                                            updated_count += 1

                                        # 从未匹配记录列表中移除这条记录（因为已经从无效数据中恢复了）
                                        # 这样就不会再从未匹配记录中添加重复的记录
                                        # 注意：这里需要用索引来移除，因为row是Series的副本
                                        # 标记这条记录为已匹配，后续不再处理
                                        row['_matched'] = True

                                if len(matches) > 0:
                                    # 找到匹配,更新论文题目
                                    matched_count += 1
                                    for match_idx in matches.index:
                                        old_title = base_df.loc[match_idx, '课题名称']
                                        if old_title != new_title:
                                            base_df.loc[match_idx, '课题名称'] = new_title
                                            updated_count += 1
                                else:
                                    # 未找到匹配,保存到未匹配记录列表
                                    unmatched_records.append(row.to_dict())

                        class_info = f" [班级: {valid_df['文件提取的班级'].iloc[0]}]" if valid_df is not None and '文件提取的班级' in valid_df.columns and valid_df['文件提取的班级'].iloc[0] else ""
                        record_count = len(valid_df) if valid_df is not None else 0
                        print(f"✓ 已读取: {file.name}{class_info} ({record_count} 条记录)")

                except Exception as e:
                    print(f"✗ 读取失败: {file.name} - {e}")
                    Logger.error(f"读取二次提交文件失败: {file.name} - {e}")

            print(f"\n✓ 论文题目更新完成: 匹配 {matched_count} 条, 更新 {updated_count} 条题目")
            if len(unmatched_records) > 0:
                print(f"  ! 未找到匹配: {len(unmatched_records)} 条记录(将添加到基础数据库)")

        # 将未匹配的记录添加到基础数据中
        if len(unmatched_records) > 0:
            # 过滤掉已经从无效数据中恢复的记录
            unmatched_records_filtered = [r for r in unmatched_records if not r.get('_matched', False)]

            if len(unmatched_records_filtered) > 0:
                print(f"\n正在添加 {len(unmatched_records_filtered)} 条未匹配记录到基础数据库...")

                # 将未匹配记录转换为DataFrame
                unmatched_df = pd.DataFrame(unmatched_records_filtered)

                # 为每条未匹配记录添加未匹配原因
                unmatched_df['未匹配原因'] = '原始数据中不存在该学生记录'

                # 补充缺失的列
                required_columns = [
                    '课题名称', '可选范围', '所属专业', '指导教师姓名', '指导教师工号',
                    '学生姓名', '学生学号', '学生组织', '来源', '模板类型', '第二指导教师',
                    '文件提取的班级', '来源文件'
                ]

                for col in required_columns:
                    if col not in unmatched_df.columns:
                        unmatched_df[col] = ''

                # 填充可选范围
                unmatched_df['可选范围'] = '汽车工程学院'

                # 填充学生组织
                from config import DEFAULT_ORGANIZATION
                unmatched_df['学生组织'] = DEFAULT_ORGANIZATION

                # 填充模板类型
                from config import DEFAULT_TEMPLATE_TYPE
                unmatched_df['模板类型'] = DEFAULT_TEMPLATE_TYPE

                # 填充第二指导教师
                from config import DEFAULT_SECOND_TEACHER
                unmatched_df['第二指导教师'] = DEFAULT_SECOND_TEACHER

                # 保存未匹配记录用于输出（在合并前保存）
                unmatched_for_output = unmatched_df.copy()

                # 合并到基础数据
                base_df = pd.concat([base_df, unmatched_df], ignore_index=True)
                print(f"✓ 已添加 {len(unmatched_df)} 条记录到基础数据库")
                print(f"✓ 更新后基础数据库总计: {len(base_df)} 条记录")
            else:
                unmatched_for_output = pd.DataFrame()

            # 清空无效数据，避免重复处理
            base_invalid_df = pd.DataFrame()
        else:
            unmatched_for_output = pd.DataFrame()

        return base_df, base_invalid_df, unmatched_for_output

    def _read_excel_with_auto_header(self, file):
        """读取Excel文件,自动检测表头行"""
        try:
            # 根据文件扩展名选择引擎
            if file.suffix.lower() == '.xls':
                # 旧的.xls格式，尝试使用xlrd引擎
                try:
                    df = pd.read_excel(file, engine='xlrd', sheet_name=0, header=0)
                except ImportError:
                    # xlrd未安装，尝试使用openpyxl（可能失败）
                    print(f"  警告: 需要安装xlrd库来读取.xls文件: pip install xlrd")
                    return None
                except Exception as e:
                    print(f"  警告: 无法读取.xls文件 {file.name}: {e}")
                    return None
            else:
                # 新的.xlsx格式，使用openpyxl
                df = pd.read_excel(file, engine='openpyxl', sheet_name=0, header=0)

            # 检查是否成功读取到有效的列名
            has_valid_columns = False
            if len(df.columns) > 5:  # 至少要有6列以上
                valid_col_count = 0
                for col in df.columns:
                    col_str = str(col)
                    if any(keyword in col_str for keyword in ['课题', '题目', '学生', '教师', '学号', '姓名']):
                        valid_col_count += 1

                if valid_col_count >= 3:
                    has_valid_columns = True

            # 如果第一行不是表头,尝试后面的行
            if not has_valid_columns:
                for header_row in [1, 2, 3]:
                    try:
                        # 根据文件扩展名选择引擎
                        if file.suffix.lower() == '.xls':
                            df_test = pd.read_excel(file, engine='xlrd', sheet_name=0, header=header_row)
                        else:
                            df_test = pd.read_excel(file, engine='openpyxl', sheet_name=0, header=header_row)

                        valid_col_count = 0
                        for col in df_test.columns:
                            col_str = str(col)
                            if any(keyword in col_str for keyword in ['课题', '题目', '学生', '教师', '学号', '姓名']):
                                valid_col_count += 1

                        if valid_col_count >= 3:
                            df = df_test
                            break
                    except:
                        continue

            return df

        except Exception as e:
            # 尝试手动读取（仅用于.xlsx文件）
            if file.suffix.lower() == '.xls':
                # .xls文件不支持手动读取，直接返回None
                return None

            try:
                from openpyxl import load_workbook
                import warnings
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                wb = load_workbook(file, data_only=True, read_only=True)
                ws = wb.active

                # 查找真正的表头行
                header_row_idx = None
                for row_idx in range(min(10, ws.max_row)):
                    row_data = []
                    for cell in ws[row_idx + 1]:
                        row_data.append(cell.value if cell.value is not None else '')

                    col_str = ' '.join([str(x) for x in row_data])
                    if any(keyword in col_str for keyword in ['课题', '题目', '学生姓名', '指导教师']):
                        header_row_idx = row_idx
                        break

                # 提取数据
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)

                if data and len(data) > header_row_idx + 1:
                    df = pd.DataFrame(data[header_row_idx + 1:], columns=data[header_row_idx])
                else:
                    df = None

                wb.close()
                return df
            except:
                return None

    def merge_student_data(self, df):
        """合并同一学生的多条记录，以二次提交为准，从原始数据补充缺失信息

        判断同一学生的标准：
        - 标准格式：学生姓名 AND 学号相同
        - 简化格式（二次提交）：学生姓名 AND 班级/所属专业相同
        """
        # 确保有学生姓名列
        if '学生姓名' not in df.columns:
            return df

        # 查找学生学号列
        student_id_col = None
        for col in df.columns:
            if '学号' in str(col):
                student_id_col = col
                break

        # 检查是否有简化格式的数据
        has_simplified = 'is_simplified_format' in df.columns and df['is_simplified_format'].any()

        # 创建唯一标识符
        # 策略：为无学号的记录尝试从同名同导师的记录中"借用"学号
        # 步骤：
        # 1. 先找出所有有学号的记录，用学号作为键
        # 2. 对于无学号的记录，尝试从同名同导师的记录中找到学号
        # 3. 如果找到了，用学号作为键；否则，用导师作为键

        # 创建一个字典，存储（姓名，导师）→ 学号的映射
        name_teacher_to_id = {}
        if student_id_col and student_id_col in df.columns:
            for idx, row in df.iterrows():
                student_id = row.get(student_id_col)
                teacher = row.get('指导教师姓名') if '指导教师姓名' in df.columns else None
                name = str(row['学生姓名']).strip()  # 去除空格

                # 如果这条记录有学号，记录下来
                if pd.notna(student_id) and str(student_id).strip() not in ['', 'nan']:
                    if pd.notna(teacher) and str(teacher).strip() not in ['', 'nan']:
                        teacher_clean = str(teacher).strip()  # 去除空格
                        key = (name, teacher_clean)
                        # 只记录第一个学号
                        if key not in name_teacher_to_id:
                            name_teacher_to_id[key] = str(student_id)

        def get_unified_key(row):
            """生成统一的匹配键"""
            name = str(row['学生姓名']).strip()  # 去除空格
            teacher = row.get('指导教师姓名') if '指导教师姓名' in df.columns else None
            teacher_clean = str(teacher).strip() if pd.notna(teacher) else ''

            # 1. 如果有学号，直接使用
            if student_id_col and student_id_col in df.columns:
                student_id = row.get(student_id_col)
                if pd.notna(student_id) and str(student_id).strip() not in ['', 'nan']:
                    return name + '|||' + str(student_id)

            # 2. 如果没有学号，尝试从字典中借用
            if teacher_clean and teacher_clean not in ['', 'nan']:
                key = (name, teacher_clean)
                if key in name_teacher_to_id:
                    # 找到了同名同导师的学号，使用它
                    return name + '|||' + name_teacher_to_id[key]

                # 没找到，使用导师作为键
                return name + '|||TEACHER|||' + teacher_clean

            # 3. 如果什么都没有，只用姓名
            return name

        df['_student_key'] = df.apply(get_unified_key, axis=1)

        # 按唯一标识符分组
        grouped = df.groupby('_student_key')
        merged_indices = []
        duplicate_count = 0

        for student_key, group in grouped:
            if len(group) == 1:
                # 只有一条记录，直接保留
                merged_indices.append(group.index[0])
            else:
                # 多条记录，需要合并（同一学生在二次提交和原始数据中都存在）
                duplicate_count += 1
                # 优先使用 is_resubmit=True 的记录
                resubmit_records = group[group['is_resubmit'] == True]
                original_records = group[group['is_resubmit'] == False]

                if len(resubmit_records) > 0:
                    # 以二次提交记录为基础
                    base_idx = resubmit_records.index[0]
                    # 如果有多条二次提交记录，使用第一条
                    # 从原始数据补充缺失信息（课题名称除外）
                    if len(original_records) > 0:
                        # 就地修改基础记录
                        for col in df.columns:
                            # 跳过课题名称列 - 始终使用二次提交的课题名称
                            if '课题' in str(col) or '题目' in str(col) or '标题' in str(col):
                                continue
                            # 跳过临时列和内部标记列
                            if col.startswith('_') or col in ['is_resubmit', 'is_simplified_format']:
                                continue

                            base_val = df.at[base_idx, col]
                            # 判断基础值是否为空
                            is_empty = (
                                pd.isna(base_val) or
                                str(base_val).strip() == '' or
                                str(base_val).lower() == 'nan'
                            )
                            if is_empty:
                                # 从原始数据获取值
                                orig_idx = original_records.index[0]
                                orig_val = df.at[orig_idx, col]
                                # 确保补充的值不为空
                                if not pd.isna(orig_val) and str(orig_val).strip() != '' and str(orig_val).lower() != 'nan':
                                    df.at[base_idx, col] = orig_val
                    merged_indices.append(base_idx)
                else:
                    # 没有二次提交记录（理论上不应该发生），使用第一条原始记录
                    merged_indices.append(group.index[0])

        # 删除临时列
        if '_student_key' in df.columns:
            df = df.drop(columns=['_student_key'])

        # 筛选合并后的记录
        merged_df = df.loc[merged_indices].reset_index(drop=True)

        if duplicate_count > 0:
            print(f"  ✓ 合并了 {duplicate_count} 个学生的重复记录")

        return merged_df

    def is_valid_title(self, title):
        """判断题目是否有效（非空、非待定等无意义内容）"""
        if pd.isna(title) or title == '' or str(title).strip() == '':
            return False

        title_str = str(title).strip()

        # 检查无意义的关键词（使用词边界匹配，避免误匹配正常词汇）
        # 通过正则表达式确保匹配的是完整词语或独立状态
        import re

        # 定义无效模式：完全匹配、前后有空格或特殊符号、纯英文/数字状态
        invalid_patterns = [
            r'^待定$',          # 完全匹配"待定"
            r'^暂定$',          # 完全匹配"暂定"
            r'^待填$',          # 完全匹配"待填"
            r'^\s*无\s*$',      # 仅包含"无"
            r'^空题?$',         # "空"或"空题"
            r'^待选$',          # 完全匹配"待选"
            r'^未定$',          # 完全匹配"未定"
            r'^nan$',           # 仅"nan"
            r'^None$',          # 仅"None"
            r'^入伍$',          # 入伍（已参军）
        ]

        for pattern in invalid_patterns:
            if re.search(pattern, title_str, re.IGNORECASE):
                return False

        # 额外检查：去除书名号《》后，如果内容过短也可能是无效
        content = title_str.replace('《', '').replace('》', '').strip()
        if len(content) <= 2:
            return False

        # 检查是否为纯符号或无意义字符
        if not re.search(r'[\u4e00-\u9fa5a-zA-Z]', content):  # 不包含中文或英文
            return False

        return True

    def remove_all_spaces(self, value):
        """去除字符串中的所有空格和空白字符，包括不间断空格"""
        return DataCleaner.remove_all_spaces(value)

    def normalize_major(self, major_value, source_class):
        """规范化所属专业字段，只返回两个标准值之一"""
        return FileHelper.normalize_major(major_value, source_class)

    def normalize_columns(self, df):
        """标准化列名，提取所需信息"""
        # 原始文件的完整列名（11列）
        original_columns = [
            '课题名称', '可选范围', '所属专业', '指导教师姓名', '指导教师工号',
            '学生姓名', '学生学号', '学生组织', '来源', '模板类型', '第二指导教师'
        ]

        # 查找各列（使用模糊匹配）
        column_mapping = {}

        # 查找课题名称列
        for col in df.columns:
            if '课题' in str(col) or '题目' in str(col) or '标题' in str(col):
                column_mapping['课题名称'] = col
                break

        # 查找其他列 - 使用模糊匹配(注意顺序很重要,更具体的要放在前面)
        for col in df.columns:
            col_str = str(col).strip()

            # 可选范围
            if '可选范围' in col_str:
                column_mapping['可选范围'] = col
            # 所属专业
            elif '所属专业' in col_str:
                column_mapping['所属专业'] = col
            # 指导教师姓名 (必须优先匹配,避免被指导教师工号覆盖)
            elif '指导教师姓名' in col_str or col_str == '指导教师':
                column_mapping['指导教师姓名'] = col
            # 指导教师工号
            elif ('指导教师工号' in col_str or '教师工号' in col_str or ('工号' in col_str and '指导' in col_str)):
                column_mapping['指导教师工号'] = col
            # 学生姓名 (匹配"学生姓名"或"姓名")
            elif '学生姓名' in col_str or (col_str == '姓名'):
                column_mapping['学生姓名'] = col
            # 学生学号
            elif '学生学号' in col_str:
                column_mapping['学生学号'] = col
            # 学生组织
            elif '学生组织' in col_str:
                column_mapping['学生组织'] = col
            # 来源
            elif '来源' in col_str and '学生' not in col_str:  # 避免匹配到"学生来源"
                column_mapping['来源'] = col
            # 模板类型
            elif '模板类型' in col_str:
                column_mapping['模板类型'] = col
            # 第二指导教师
            elif '第二指导教师' in col_str:
                column_mapping['第二指导教师'] = col

        # 创建标准化的数据，包含所有原始列
        result = pd.DataFrame()

        # 填充所有列（如果找不到则填充空值）
        for orig_col in original_columns:
            if orig_col in column_mapping:
                result[orig_col] = df[column_mapping[orig_col]].astype(str)
            else:
                result[orig_col] = ''

        # 立即去除课题名称中的所有空格（在所有处理之前）
        result['课题名称'] = result['课题名称'].apply(self.remove_all_spaces)

        # 立即规范化可选范围字段（去除首尾空格，替换nan为空字符串）
        if '可选范围' in result.columns:
            def clean_optional_range(val):
                val_str = str(val).strip()
                return '' if val_str in ['nan', 'None', 'NA'] else val_str
            result['可选范围'] = result['可选范围'].apply(clean_optional_range)

        # 添加辅助列
        source_class = ''
        if 'source_class' in df.columns:
            # 使用从文件名提取的班级
            source_class = df['source_class'].astype(str)
            result['文件提取的班级'] = source_class
        else:
            result['文件提取的班级'] = ''

        result['来源文件'] = df['source_file'].astype(str) if 'source_file' in df.columns else ''
        result['原始索引'] = df.index

        # 保留内部标记列（用于后续处理）
        if 'is_resubmit' in df.columns:
            result['is_resubmit'] = df['is_resubmit']
        if 'is_simplified_format' in df.columns:
            result['is_simplified_format'] = df['is_simplified_format']
        if '班级' in df.columns:
            result['班级'] = df['班级']

        # 规范化所属专业字段为两个标准值之一
        result['所属专业'] = result.apply(
            lambda row: self.normalize_major(row['所属专业'], row['文件提取的班级']),
            axis=1
        )

        # 使用课题名称列进行有效性判断（此时课题名称已无空格）
        valid_mask = result['课题名称'].apply(self.is_valid_title)
        valid_df = result[valid_mask].reset_index(drop=True)
        invalid_df = result[~valid_mask].reset_index(drop=True)

        return valid_df, invalid_df

    def calculate_similarity(self, titles):
        """计算题目之间的相似度"""
        # 使用jieba分词
        print("正在分词和计算相似度...")

        # 使用列表推导式和并行处理提升性能
        text_split = [' '.join(jieba.lcut(str(t))) for t in titles]

        # TF-IDF向量化
        vectorizer = TfidfVectorizer()
        tfidf_matrix = vectorizer.fit_transform(text_split)

        # 计算余弦相似度
        similarity_matrix = cosine_similarity(tfidf_matrix, tfidf_matrix)
        return similarity_matrix

    def find_similar_pairs(self, df, similarity_matrix):
        """找出相似度超过阈值的题目对"""
        similar_pairs = []
        n = len(df)

        # 预先清理学生姓名和导师姓名,避免在循环中重复操作
        df_clean = df.copy()
        df_clean['学生姓名_clean'] = df_clean['学生姓名'].apply(lambda x: str(x).strip())
        df_clean['指导教师姓名_clean'] = df_clean['指导教师姓名'].apply(lambda x: str(x).strip())

        # 使用numpy找出所有超过阈值的索引对
        import numpy as np
        rows, cols = np.where(similarity_matrix >= self.threshold)
        # 只保留上三角矩阵(i < j)
        candidate_pairs = [(i, j) for i, j in zip(rows, cols) if i < j]

        for i, j in candidate_pairs:
            similarity = similarity_matrix[i][j]

            # 获取学生和导师信息
            student_a_clean = df_clean.iloc[i]['学生姓名_clean']
            student_b_clean = df_clean.iloc[j]['学生姓名_clean']
            teacher_a_clean = df_clean.iloc[i]['指导教师姓名_clean']
            teacher_b_clean = df_clean.iloc[j]['指导教师姓名_clean']

            # 跳过同一学生的自己比较
            is_same_student = False

            if student_a_clean == student_b_clean:
                # 学生姓名相同，进一步检查
                # 1. 如果导师相同，肯定是同一学生
                if teacher_a_clean == teacher_b_clean:
                    is_same_student = True
                # 2. 如果都有学号且学号相同，也是同一学生
                elif ('学生学号' in df.columns and
                      pd.notna(df.iloc[i]['学生学号']) and
                      pd.notna(df.iloc[j]['学生学号']) and
                      str(df.iloc[i]['学生学号']).strip() == str(df.iloc[j]['学生学号']).strip() and
                      str(df.iloc[i]['学生学号']).strip() not in ['nan', '']):
                    is_same_student = True

            if is_same_student:
                continue

            # 优先使用文件提取的班级，其次使用所属专业
            class_a = df.iloc[i]['文件提取的班级'] if df.iloc[i]['文件提取的班级'] else df.iloc[i]['所属专业']
            class_b = df.iloc[j]['文件提取的班级'] if df.iloc[j]['文件提取的班级'] else df.iloc[j]['所属专业']

            similar_pairs.append({
                '题目A': df.iloc[i]['课题名称'],
                '题目B': df.iloc[j]['课题名称'],
                '学生A': df_clean.iloc[i]['学生姓名'],
                '学生B': df_clean.iloc[j]['学生姓名'],
                '导师A': df_clean.iloc[i]['指导教师姓名'],
                '导师B': df_clean.iloc[j]['指导教师姓名'],
                '班级A': class_a,
                '班级B': class_b,
                '相似度': f"{similarity:.2%}",
                '相似度数值': similarity
            })

        return sorted(similar_pairs, key=lambda x: x['相似度数值'], reverse=True)

    def format_excel_sheet(self, ws, header_row=True):
        """格式化Excel工作表"""
        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 定义表头样式
        header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=HEADER_FONT_SIZE)
        header_fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 定义数据单元格样式
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # 应用格式到所有行
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                # 应用边框
                cell.border = thin_border

                # 第一行是表头
                if row_idx == 1 and header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    cell.alignment = cell_alignment
                    # 自动调整行高
                    if row_idx > 1:
                        try:
                            value = str(cell.value) if cell.value else ""
                            # 根据内容长度估算行高（每行约15个字符）
                            lines = max(1, (len(value) + 20) // 25)
                            ws.row_dimensions[row_idx].height = max(15, lines * 15)
                        except:
                            ws.row_dimensions[row_idx].height = 15

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        # 计算字符串长度（中文字符按2个长度计算）
                        value = str(cell.value)
                        length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in value)
                        if length > max_length:
                            max_length = length
                except:
                    pass

            # 设置列宽，限制在合理范围内
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            if adjusted_width < 12:
                adjusted_width = 12  # 最小宽度12
            ws.column_dimensions[column_letter].width = adjusted_width

    def highlight_similar_rows(self, ws):
        """为相似题目工作表的高相似度行添加颜色标记"""
        # 定义颜色填充
        high_fill = PatternFill(start_color=COLOR_HIGH_SIMILARITY, end_color=COLOR_HIGH_SIMILARITY, fill_type='solid')  # 红色
        medium_fill = PatternFill(start_color=COLOR_MEDIUM_SIMILARITY, end_color=COLOR_MEDIUM_SIMILARITY, fill_type='solid')  # 黄色

        # 从第2行开始（第1行是表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            # 相似度在最后一列
            similarity_cell = row[-1]
            if similarity_cell.value:
                try:
                    # 提取相似度数值（去掉百分号）
                    sim_str = str(similarity_cell.value).replace('%', '')
                    sim_value = float(sim_str) / 100

                    # 根据相似度设置背景色
                    if sim_value >= 0.85:
                        for cell in row:
                            cell.fill = high_fill
                    elif sim_value >= 0.80:
                        for cell in row:
                            cell.fill = medium_fill
                except:
                    pass

    def highlight_invalid_rows(self, ws):
        """为全部数据工作表的无效题目行添加颜色标记"""
        # 定义无效数据的灰色背景
        invalid_fill = PatternFill(start_color=COLOR_INVALID_ROW, end_color=COLOR_INVALID_ROW, fill_type='solid')  # 浅灰色

        # 从第2行开始（第1行是表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            # 检查最后一列的"数据状态"
            status_cell = row[-1]
            if status_cell.value and str(status_cell.value) == '无效':
                for cell in row:
                    cell.fill = invalid_fill

    def highlight_all_data_rows(self, ws, similar_pairs, all_data_display):
        """为全部数据工作表的相似题目和无效题目添加颜色标记"""
        # 定义颜色填充
        invalid_fill = PatternFill(start_color=COLOR_INVALID_ROW, end_color=COLOR_INVALID_ROW, fill_type='solid')  # 浅灰色 - 无效题目
        similar_fill = PatternFill(start_color=COLOR_SIMILAR_CELL, end_color=COLOR_SIMILAR_CELL, fill_type='solid')  # 黄色 - 相似题目

        # 收集所有相似题目的课题名称
        similar_titles = set()
        for pair in similar_pairs:
            similar_titles.add(str(pair['题目A']))
            similar_titles.add(str(pair['题目B']))

        # 从第2行开始（第1行是表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            # 获取该行的课题名称（第2列，索引为1）
            title_cell = row[1]  # 序号是第0列，课题名称是第1列

            # 优先标记无效题目（灰色）
            status_cell = row[-1]  # 最后一列是数据状态
            is_invalid = status_cell.value and str(status_cell.value) == '无效'

            # 检查是否为相似题目
            is_similar = title_cell.value and str(title_cell.value) in similar_titles

            # 应用颜色
            if is_invalid:
                # 无效题目 - 灰色
                for cell in row:
                    cell.fill = invalid_fill
            elif is_similar:
                # 相似题目 - 黄色
                for cell in row:
                    cell.fill = similar_fill

    def run(self):
        """执行查重"""
        print("=" * 60)
        print("论文选题查重系统")
        print("=" * 60)

        # 检查并创建数据文件夹
        if not self.data_dir.exists():
            self.data_dir.mkdir(parents=True, exist_ok=True)
            print("\n提示：已自动创建「原始数据」文件夹")
            print("请将待检测的 Excel 文件放入「原始数据」文件夹后，再运行程序")
            return

        if not self.resubmit_dir.exists():
            self.resubmit_dir.mkdir(parents=True, exist_ok=True)

        if not self.output_dir.exists():
            self.output_dir.mkdir(parents=True, exist_ok=True)

        # 1. 读取数据
        valid_df, invalid_df, unmatched_records = self.read_excel_files()

        # 输出无效题目信息
        if len(invalid_df) > 0:
            print(f"发现 {len(invalid_df)} 条无效题目记录（空题、待定等）")

        # 2. 数据已经在read阶段完成合并和规范化
        print(f"有效记录数: {len(valid_df)}")

        # 3. 计算相似度（只对有效题目）
        similarity_matrix = self.calculate_similarity(valid_df['课题名称'].tolist())

        # 4. 找出相似题目对
        similar_pairs = self.find_similar_pairs(valid_df, similarity_matrix)

        # 5. 输出结果到多个工作表
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = self.output_dir / f"查重结果_{timestamp}.xlsx"

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 工作表1：相似题目对
            if similar_pairs:
                result_df = pd.DataFrame(similar_pairs)
                result_df = result_df.drop(columns=['相似度数值'])
                result_df.to_excel(writer, sheet_name='相似题目', index=False)
            else:
                empty_df = pd.DataFrame(columns=['题目A', '题目B', '学生A', '学生B', '导师A', '导师B', '班级A', '班级B', '相似度'])
                empty_df.to_excel(writer, sheet_name='相似题目', index=False)

            # 工作表2：无效题目记录
            if len(invalid_df) > 0:
                # 选择主要列显示
                invalid_display = invalid_df[['课题名称', '学生姓名', '指导教师姓名', '所属专业', '文件提取的班级', '来源文件']].copy()
                # 合并班级列（优先显示文件提取的班级）
                invalid_display['班级'] = invalid_display['文件提取的班级'].fillna(invalid_display['所属专业'])
                invalid_display = invalid_display[['课题名称', '学生姓名', '指导教师姓名', '班级', '来源文件']]
                invalid_display.columns = ['课题名称', '学生姓名', '指导教师姓名', '班级', '来源文件']
                invalid_display.to_excel(writer, sheet_name='无效题目', index=False)
            else:
                empty_invalid_df = pd.DataFrame(columns=['课题名称', '学生姓名', '指导教师姓名', '班级', '来源文件'])
                empty_invalid_df.to_excel(writer, sheet_name='无效题目', index=False)

            # 工作表3：所有原始数据（合并所有列，包括有效和无效）
            # 合并有效和无效数据
            all_data_with_invalid = pd.concat([valid_df, invalid_df], ignore_index=True)

            # 选择所有原始数据列
            all_columns = [
                '课题名称', '可选范围', '所属专业', '指导教师姓名', '指导教师工号',
                '学生姓名', '学生学号', '学生组织', '来源', '模板类型', '第二指导教师',
                '文件提取的班级', '来源文件'
            ]

            # 确保所有列都存在
            all_data_display = all_data_with_invalid.copy()
            for col in all_columns:
                if col not in all_data_display.columns:
                    all_data_display[col] = ''

            all_data_display = all_data_display[all_columns]

            # 数据规范化
            import random
            random.seed(RANDOM_SEED)  # 固定随机种子，确保可重复

            # 注意：课题名称中的空格已在 normalize_columns 阶段去除

            # 1. 可选范围统一为"汽车工程学院"（去除空格、标点、替换nan）
            def normalize_optional_range(value):
                """规范化可选范围字段"""
                return DataCleaner.normalize_optional_range(value)

            all_data_display['可选范围'] = all_data_display['可选范围'].apply(normalize_optional_range)

            # 2. 学生组织统一为配置文件中的默认组织
            all_data_display['学生组织'] = DEFAULT_ORGANIZATION

            # 3. 来源为空则随机填写"学生自选"或"教师指定"
            def fill_source(value):
                if pd.isna(value) or str(value).strip() == '' or str(value) == 'nan':
                    return random.choice(['学生自选', '教师指定'])
                return value

            all_data_display['来源'] = all_data_display['来源'].apply(fill_source)

            # 4. 模板类型统一为配置文件中的默认类型
            all_data_display['模板类型'] = DEFAULT_TEMPLATE_TYPE

            # 5. 第二指导教师为空则填写配置文件中的默认值
            def fill_second_teacher(value):
                if pd.isna(value) or str(value).strip() == '' or str(value) == 'nan':
                    return DEFAULT_SECOND_TEACHER
                return value

            all_data_display['第二指导教师'] = all_data_display['第二指导教师'].apply(fill_second_teacher)

            # 6. 指导教师工号：规范化（去除小数点，空值填写"无"）
            def clean_teacher_id(value):
                """规范化教师工号"""
                return DataCleaner.normalize_teacher_id(value)

            all_data_display['指导教师工号'] = all_data_display['指导教师工号'].apply(clean_teacher_id)

            # 7. 学生学号：规范化（空值填写"无"）
            def normalize_student_id(value):
                """规范化学生学号"""
                return DataCleaner.normalize_student_id(value)

            all_data_display['学生学号'] = all_data_display['学生学号'].apply(normalize_student_id)

            # 6. 添加数据有效性标注列
            def mark_validity(title):
                """标记题目是否有效"""
                if self.is_valid_title(title):
                    return '有效'
                else:
                    return '无效'

            all_data_display.insert(len(all_data_display.columns), '数据状态', all_data_display['课题名称'].apply(mark_validity))

            # 添加序号列
            all_data_display.insert(0, '序号', range(1, len(all_data_display) + 1))

            # 重命名列以更清晰
            all_data_display.columns = [
                '序号', '课题名称', '可选范围', '所属专业', '指导教师姓名', '指导教师工号',
                '学生姓名', '学生学号', '学生组织', '来源', '模板类型', '第二指导教师',
                '班级(文件名)', '来源文件', '数据状态'
            ]

            # 按班级和学号排序
            # 首先尝试按学生学号（数字）排序，如果学号无法转换为数字则按字符串排序
            def safe_student_id_sort(id_value):
                """安全地提取学号中的数字部分用于排序"""
                try:
                    # 尝试提取学号中的数字部分
                    id_str = str(id_value)
                    # 提取所有数字
                    import re
                    numbers = re.findall(r'\d+', id_str)
                    if numbers:
                        # 返回第一个数字序列作为排序键
                        return int(numbers[0])
                    return 0
                except:
                    return 0

            # 创建排序列
            all_data_display['_sort_class'] = all_data_display['班级(文件名)'].fillna('')
            all_data_display['_sort_id'] = all_data_display['学生学号'].apply(safe_student_id_sort)

            # 先按班级排序，再按学号排序
            all_data_display = all_data_display.sort_values(
                by=['_sort_class', '_sort_id'],
                ascending=[True, True]
            )

            # 删除临时排序列
            all_data_display = all_data_display.drop(columns=['_sort_class', '_sort_id'])

            # 重新生成序号
            all_data_display['序号'] = range(1, len(all_data_display) + 1)

            all_data_display.to_excel(writer, sheet_name='全部数据', index=False)

            # 工作表4：未匹配数据
            if unmatched_records is not None and len(unmatched_records) > 0:
                # 将未匹配记录转换为DataFrame
                unmatched_df = pd.DataFrame(unmatched_records)

                # 选择主要列显示（包含未匹配原因）
                unmatched_columns = [
                    '课题名称', '学生姓名', '学生学号', '指导教师姓名', '指导教师工号',
                    '所属专业', '文件提取的班级', '来源文件', '未匹配原因'
                ]

                # 确保所有列都存在
                for col in unmatched_columns:
                    if col not in unmatched_df.columns:
                        unmatched_df[col] = ''

                unmatched_display = unmatched_df[unmatched_columns]

                # 重命名列以更清晰
                unmatched_display.columns = [
                    '课题名称', '学生姓名', '学生学号', '指导教师姓名', '指导教师工号',
                    '所属专业', '班级', '来源文件', '未匹配原因'
                ]

                unmatched_display.to_excel(writer, sheet_name='未匹配数据', index=False)
            else:
                # 如果没有未匹配记录，创建空表（包含未匹配原因列）
                empty_unmatched_df = pd.DataFrame(columns=[
                    '课题名称', '学生姓名', '学生学号', '指导教师姓名', '指导教师工号',
                    '所属专业', '班级', '来源文件', '未匹配原因'
                ])
                empty_unmatched_df.to_excel(writer, sheet_name='未匹配数据', index=False)

        # 格式化Excel文件
        print("正在优化Excel格式...")
        wb = load_workbook(output_file)

        # 格式化工作表1：相似题目
        if '相似题目' in wb.sheetnames:
            ws_similar = wb['相似题目']
            self.format_excel_sheet(ws_similar)
            self.highlight_similar_rows(ws_similar)
            # 冻结首行
            ws_similar.freeze_panes = 'A2'

        # 格式化工作表2：无效题目
        if '无效题目' in wb.sheetnames:
            ws_invalid = wb['无效题目']
            self.format_excel_sheet(ws_invalid)
            # 冻结首行
            ws_invalid.freeze_panes = 'A2'

        # 格式化工作表3：全部数据
        if '全部数据' in wb.sheetnames:
            ws_all = wb['全部数据']
            self.format_excel_sheet(ws_all)
            # 标记相似题目行和无效题目行
            self.highlight_all_data_rows(ws_all, similar_pairs, all_data_display)
            # 冻结首行
            ws_all.freeze_panes = 'A2'

        # 保存格式化后的文件
        wb.save(output_file)

        if similar_pairs:
            print(f"\n发现 {len(similar_pairs)} 对相似题目 (阈值: {self.threshold:.0%})")
            print(f"结果已保存到: {output_file}")

            # 显示前10对
            print("\n相似度最高的前10对:")
            print("-" * 60)
            for i, pair in enumerate(similar_pairs[:10], 1):
                print(f"\n{i}. 相似度: {pair['相似度']}")
                print(f"   题目A: {pair['题目A']}")
                print(f"   题目B: {pair['题目B']}")
                print(f"   学生A: {pair['学生A']} | 导师A: {pair['导师A']} | 班级A: {pair['班级A']}")
                print(f"   学生B: {pair['学生B']} | 导师B: {pair['导师B']} | 班级B: {pair['班级B']}")
        else:
            print(f"\n未发现相似题目 (阈值: {self.threshold:.0%})")
            print(f"结果已保存到: {output_file}")

        print("\n查重完成!")
        print(f"有效题目数: {len(valid_df)}")
        print(f"无效题目数: {len(invalid_df)}")
        print(f"相似题目对: {len(similar_pairs)}")



def main():
    """主函数"""
    checker = ThesisChecker(
        data_dir="原始数据",
        resubmit_dir="二次提交",
        output_dir="查重结果",
        threshold=0.75  # 可调整相似度阈值
    )
    checker.run()


if __name__ == "__main__":
    main()
